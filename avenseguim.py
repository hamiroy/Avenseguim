import os
import openpyxl

def avensegium(directory, keyword):
    keyword_hits = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):
                file_path = os.path.join(root, file)
                try:
                    workbook = openpyxl.load_workbook(file_path, data_only=True)
                    for sheet in workbook.sheetnames:
                        worksheet = workbook[sheet]
                        for row in worksheet.iter_rows(values_only=True):
                            for cell in row:
                                if cell and keyword in str(cell):
                                    keyword_hits.append({'file': file_path, 'worksheet': sheet})
                                    break
                except Exception as e:
                    print(f"Error processing {file_path}: {e}")
    return keyword_hits

if __name__ == "__main__":
    directory = r'C:\Users\F3690\OneDrive\001-chronicle\#SharedFolder\New Charm Management Limited - Mount Austin Estate'
    keyword = 'Final Completion Checklist'
    hits = avensegium(directory, keyword)
    print("Keyword hits:")
    for hit in hits:
        print(f"File: {hit['file']}, Worksheet: {hit['worksheet']}")
